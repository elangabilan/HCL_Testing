from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
import time

# Load Excel
wb = load_workbook('data.xlsx')
print("Loaded Excel file")
sheet = wb.active

# URL
url = 'https://practicetestautomation.com/practice-test-login/'

# List of browser drivers
browsers = [
    ("Chrome", lambda: webdriver.Chrome(service=ChromeService())),
    ("Firefox", lambda: webdriver.Firefox(service=FirefoxService())),
    ("Edge", lambda: webdriver.Edge(service=EdgeService()))
]

# Loop over each browser
for browser_name, browser_func in browsers:
    print(f"\nRunning in {browser_name}...")
    driver = browser_func()
    driver.maximize_window()
    driver.get(url)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, email, password = row
        driver.find_element(By.XPATH, '//*[@id="username"]').clear()
        driver.find_element(By.XPATH, '//*[@id="username"]').send_keys(email)
        driver.find_element(By.XPATH, '//*[@id="password"]').clear()
        driver.find_element(By.XPATH, '//*[@id="password"]').send_keys(password)
        driver.find_element(By.XPATH, '//*[@id="submit"]').click()
        print(f"Completed for: {name}")

        time.sleep(2)
        driver.get(url)

    print(f"Completed tests in {browser_name}. Closing browser.")
    driver.quit()

print("\nAll browser tests completed successfully.")
